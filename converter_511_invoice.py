from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


PDF_COLS = [
    "Line",
    "Purchase Order",
    "Customer Item",
    "5.11 Item",
    "Description",
    "Color",
    "Size",
    "Dim",
    "Qty",
    "List Unit Price",
    "Your Unit Price",
    "Amount",
]

OUTPUT_HEADERS = [
    "",
    "Purchase Order",
    "Customer Item",
    "L-M-N_kodas",
    "Description",
    "Size/Dim",
    "Qty",
    "Your Unit Price",
    "Amount",
    "",
    "",
    "5",
    "5.11 Item",
    "Color",
    "Size",
    "Dim",
]


@dataclass
class InvoiceHeader:
    invoice_number: str
    invoice_date: str
    customer_po: str
    invoice_total: str


@dataclass
class ConversionResult:
    output_path: Path
    invoice: InvoiceHeader
    pdf_row_count: int
    excel_row_count: int
    qty_sum: int
    amount_sum: str
    validations: List[str]


class InvoiceConversionError(Exception):
    pass


def _cluster(values: List[float], tolerance: float = 1.2) -> List[float]:
    values = sorted(values)
    clusters: List[List[float]] = []
    for value in values:
        if not clusters or abs(value - clusters[-1][-1]) > tolerance:
            clusters.append([value])
        else:
            clusters[-1].append(value)
    return [sum(c) / len(c) for c in clusters]


def _money_to_decimal(value: str) -> Decimal:
    try:
        return Decimal(value.replace(",", "").strip())
    except (InvalidOperation, AttributeError) as exc:
        raise InvoiceConversionError(f"Ne skaitinė pinigų reikšmė: {value!r}") from exc


def _int_value(value: str) -> int:
    try:
        return int(str(value).strip())
    except ValueError as exc:
        raise InvoiceConversionError(f"Ne skaitinė Qty reikšmė: {value!r}") from exc


def _extract_header(full_text: str) -> InvoiceHeader:
    invoice_match = re.search(r"\b(SE\.IN-\d+)\b", full_text)
    if not invoice_match:
        raise InvoiceConversionError("Nepavyko rasti Invoice Number.")

    invoice_number = invoice_match.group(1)

    # The invoice header table contains:
    # SE.IN-00244072 REG-P01 ...
    # 6/22/2026 ...
    # Customer PO may contain spaces, e.g. "Special price".
    # It sits between Invoice Number and Customer # (usually SWE80117).
    header_line_match = re.search(rf"\b{re.escape(invoice_number)}\s+(.+?)\s+SWE\d+\b", full_text)
    if not header_line_match:
        raise InvoiceConversionError("Nepavyko patikimai rasti Customer PO šalia Invoice Number.")
    customer_po = re.sub(r"\s+", " ", header_line_match.group(1)).strip()

    # Invoice date is the first date immediately following the invoice header line.
    after_invoice = full_text[invoice_match.end(): invoice_match.end() + 400]
    date_match = re.search(r"\b\d{1,2}/\d{1,2}/\d{4}\b", after_invoice)
    if not date_match:
        raise InvoiceConversionError("Nepavyko patikimai rasti Invoice Date.")
    invoice_date = date_match.group(0)

    total_match = re.search(r"\bINVOICE\s+TOTAL\s+([0-9,]+\.\d{2})\b", full_text)
    if not total_match:
        raise InvoiceConversionError("Nepavyko rasti INVOICE TOTAL lauko.")

    invoice_total = total_match.group(1)

    # Blocking control: do not accept a total found only through subtotal/freight/vat text.
    invoice_total_context = full_text[max(0, total_match.start() - 30): total_match.end() + 30]
    if "INVOICE TOTAL" not in invoice_total_context:
        raise InvoiceConversionError("Invoice Total nepatvirtintas kaip INVOICE TOTAL laukas.")

    return InvoiceHeader(
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        customer_po=customer_po,
        invoice_total=invoice_total,
    )


def _detect_table_boundaries(page) -> Tuple[List[float], float, float]:
    """
    Detects real PDF table column boundaries from vertical ruling lines.
    This does not infer Color/Size/Dim from text token count or regex; values are assigned by PDF x-coordinates.
    """
    verticals = []
    for line in page.lines:
        if abs(line["x0"] - line["x1"]) < 0.8 and line.get("height", 0) > 5 and 150 < line.get("top", 0) < 780:
            verticals.append(line)

    if not verticals:
        raise InvoiceConversionError("Puslapyje nepavyko rasti lentelės vertikalių ribų.")

    xs = _cluster([line["x0"] for line in verticals], tolerance=1.2)
    xs = [x for x in xs if 10 < x < 590]

    candidates = []
    for x in xs:
        matching = [
            line for line in verticals
            if abs(line["x0"] - x) < 1.2
        ]
        candidates.append((x, len(matching), min(l["top"] for l in matching), max(l["bottom"] for l in matching)))

    # Only use vertical rules that repeat through the actual item rows.
    # Header-only separator lines may appear around Customer Item / 5.11 Item / Description,
    # but they do not run through the item grid. Using them shifts the whole table and corrupts Color.
    strong = sorted([x for x, count, _, _ in candidates if count >= 20])

    if len(strong) != 13:
        raise InvoiceConversionError(
            f"Nepavyko patikimai nustatyti tikrų prekių lentelės stulpelių ribų: "
            f"rasta {len(strong)}, reikia 13. Ribos: {[round(x, 1) for x in strong]}"
        )

    boundaries = strong

    expected_left = 16.0
    expected_right = 577.4
    if abs(boundaries[0] - expected_left) > 3 or abs(boundaries[-1] - expected_right) > 3:
        raise InvoiceConversionError(
            f"Lentelės ribos neatitinka tikėtino PDF grid'o: "
            f"{round(boundaries[0], 1)} - {round(boundaries[-1], 1)}"
        )

    # Get the vertical span of the detected table boundaries.
    span_lines = [
        line for line in verticals
        if any(abs(line["x0"] - x) < 1.2 for x in boundaries)
    ]
    table_top = min(line["top"] for line in span_lines)
    table_bottom = max(line["bottom"] for line in span_lines)

    if len(boundaries) != 13:
        raise InvoiceConversionError("Lentelės stulpelių ribų skaičius nėra 13.")

    return boundaries, table_top, table_bottom


def _detect_small_table_boundaries(page) -> Tuple[List[float], float, float]:
    """
    v5 small-invoice fallback.
    Used only on the first item page when the strict v4 detector fails because
    there are too few item rows to reach the strict vertical repetition threshold.
    """
    verticals = []
    for line in page.lines:
        is_vertical = abs(line["x0"] - line["x1"]) < 0.8
        if not is_vertical:
            continue

        top = line.get("top", 0)
        bottom = line.get("bottom", 0)
        height = line.get("height", 0)

        # Isolate item table below invoice metadata block.
        if height > 5 and 240 <= top < 740 and 240 < bottom < 760:
            verticals.append(line)

    if not verticals:
        raise InvoiceConversionError("Small-invoice: nepavyko rasti prekių lentelės vertikalių ribų.")

    xs = _cluster([line["x0"] for line in verticals], tolerance=1.2)
    xs = [x for x in xs if 10 < x < 590]

    candidates = []
    for x in xs:
        matching = [line for line in verticals if abs(line["x0"] - x) < 1.2]
        if matching:
            candidates.append((x, len(matching), min(l["top"] for l in matching), max(l["bottom"] for l in matching)))

    usable = sorted([x for x, count, _, _ in candidates if count >= 2])

    if len(usable) < 13:
        raise InvoiceConversionError(
            f"Small-invoice: rasta {len(usable)} lentelės ribų, reikia bent 13. "
            f"Ribos: {[round(x, 1) for x in usable]}"
        )

    possible = []
    for i in range(0, len(usable) - 12):
        group = usable[i:i + 13]
        score = abs((group[-1] - group[0]) - 561.4) + abs(group[0] - 16.0) + abs(group[-1] - 577.4)
        possible.append((score, group))

    boundaries = min(possible, key=lambda item: item[0])[1]

    if abs(boundaries[0] - 16.0) > 3 or abs(boundaries[-1] - 577.4) > 3:
        raise InvoiceConversionError(
            f"Small-invoice: lentelės ribos neatitinka patvirtinto PDF grid'o: "
            f"{round(boundaries[0], 1)} - {round(boundaries[-1], 1)}"
        )

    span_lines = [line for line in verticals if any(abs(line["x0"] - x) < 1.2 for x in boundaries)]
    table_top = min(line["top"] for line in span_lines)
    table_bottom = max(line["bottom"] for line in span_lines)

    return boundaries, table_top, table_bottom


def _chars_to_text(chars: List[Dict]) -> str:
    if not chars:
        return ""

    chars = sorted(chars, key=lambda ch: (ch["x0"], ch["top"]))
    out = ""
    prev = None

    for ch in chars:
        text = ch.get("text", "")
        if not text:
            continue

        if prev is not None:
            gap = ch["x0"] - prev["x1"]
            if gap > 1.2:
                out += " "

        out += text
        prev = ch

    return re.sub(r"\s+", " ", out).strip()


def _detect_line_rows(page, boundaries: List[float], table_top: float, table_bottom: float) -> List[Tuple[int, float, float]]:
    words = page.extract_words(x_tolerance=1, y_tolerance=2)
    line_left, line_right = boundaries[0], boundaries[1]
    rows = []

    for word in words:
        center_x = (word["x0"] + word["x1"]) / 2
        if not (line_left < center_x < line_right):
            continue

        if not re.fullmatch(r"\d+", word["text"]):
            continue

        top = word["top"]
        bottom = word["bottom"]

        # Rows must be inside the detected item table area and below the header band.
        if table_top + 15 <= top <= table_bottom + 2:
            rows.append((int(word["text"]), top, bottom))

    rows = sorted(rows, key=lambda item: item[1])

    # Remove duplicates if the PDF exposes the same line number twice.
    unique = []
    seen = set()
    for line_no, top, bottom in rows:
        key = (line_no, round(top, 1))
        if key not in seen:
            seen.add(key)
            unique.append((line_no, top, bottom))

    return unique


def _detect_line_rows_by_sequence(
    page,
    boundaries: List[float],
    expected_line: int,
    table_top: Optional[float] = None,
    table_bottom: Optional[float] = None,
) -> List[Tuple[int, float, float]]:
    """
    Detects item rows by real Line column x-position and expected consecutive invoice line numbers.
    This avoids losing short final pages where the PDF has fewer vertical ruling repetitions.
    """
    words = page.extract_words(x_tolerance=1, y_tolerance=2)
    line_left, line_right = boundaries[0], boundaries[1]

    min_top = (table_top + 15) if table_top is not None else 185
    max_top = (table_bottom + 8) if table_bottom is not None else 650

    candidates = []
    for word in words:
        center_x = (word["x0"] + word["x1"]) / 2
        if not (line_left < center_x < line_right):
            continue
        if not re.fullmatch(r"\d+", word["text"]):
            continue

        top = word["top"]
        bottom = word["bottom"]
        if min_top <= top <= max_top:
            candidates.append((int(word["text"]), top, bottom))

    candidates = sorted(candidates, key=lambda item: item[1])

    rows = []
    last_line = expected_line - 1
    for line_no, top, bottom in candidates:
        # Some invoices skip a visual line number, e.g. 44 is absent and the next PDF row is 45.
        # Do not force continuous numbering; only require strictly increasing item line numbers.
        if line_no > last_line:
            rows.append((line_no, top, bottom))
            last_line = line_no

    return rows


def _extract_table_rows(pdf_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    last_boundaries: Optional[List[float]] = None
    expected_line = 1

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            boundaries = None
            table_top = None
            table_bottom = None

            try:
                boundaries, table_top, table_bottom = _detect_table_boundaries(page)
                last_boundaries = boundaries
            except InvoiceConversionError:
                # Short continuation/final pages may not contain enough repeated vertical rules.
                # If a previous page already confirmed the grid, reuse it.
                if last_boundaries is not None:
                    boundaries = last_boundaries
                else:
                    # v5: small invoice first page fallback.
                    boundaries, table_top, table_bottom = _detect_small_table_boundaries(page)
                    last_boundaries = boundaries

            line_rows = _detect_line_rows_by_sequence(
                page=page,
                boundaries=boundaries,
                expected_line=expected_line,
                table_top=table_top,
                table_bottom=table_bottom,
            )

            if not line_rows:
                continue

            for idx, (line_no, top, bottom) in enumerate(line_rows):
                next_top = line_rows[idx + 1][1] if idx + 1 < len(line_rows) else bottom + 7
                row_top = top - 1.0
                row_bottom = min(next_top - 0.4, bottom + 3.8)

                row_chars = []
                for ch in page.chars:
                    center_y = (ch["top"] + ch["bottom"]) / 2
                    if row_top <= center_y <= row_bottom:
                        row_chars.append(ch)

                values = []
                for col_idx in range(len(PDF_COLS)):
                    x0, x1 = boundaries[col_idx], boundaries[col_idx + 1]
                    col_chars = [
                        ch for ch in row_chars
                        if x0 + 0.2 <= (ch["x0"] + ch["x1"]) / 2 <= x1 - 0.2
                    ]
                    values.append(_chars_to_text(col_chars))

                row = dict(zip(PDF_COLS, values))

                if row.get("Line", "").strip() != str(line_no):
                    raise InvoiceConversionError(
                        f"Eilutės numeris nesutampa su PDF Line stulpeliu: tikėtasi {line_no}, gauta {row.get('Line')!r}."
                    )

                if not row.get("Purchase Order") or not row.get("Customer Item") or not row.get("5.11 Item"):
                    raise InvoiceConversionError(f"Nepilnai ištraukta prekių eilutė {line_no}.")

                rows.append(row)
                expected_line = line_no + 1

    if not rows:
        raise InvoiceConversionError("Nepavyko ištraukti nė vienos prekių lentelės eilutės.")

    # Blocking control: invoice line numbers must be strictly increasing and unique.
    # Do not require continuity because some PDFs may skip a visible line number.
    actual_lines = [int(row["Line"]) for row in rows]
    if actual_lines != sorted(set(actual_lines)):
        raise InvoiceConversionError(
            f"Prekių eilučių numeriai nėra griežtai didėjantys arba yra dublikatų: "
            f"{actual_lines[:5]}...{actual_lines[-5:]}"
        )

    return rows

def _remove_trailing_abr(item: str) -> str:
    return item[:-3] if item.endswith("ABR") else item


def _color_code(color: str) -> str:
    if ":" not in color:
        raise InvoiceConversionError(f"Color reikšmėje nėra ':' simbolio: {color!r}")
    code = color.split(":", 1)[0]
    if not code.isdigit():
        raise InvoiceConversionError(f"Color kodas nėra skaitinis: {color!r}")
    return code


def _build_output_rows(pdf_rows: List[Dict[str, str]]) -> List[List]:
    output_rows = []

    for row in pdf_rows:
        item_m = _remove_trailing_abr(row["5.11 Item"].strip())
        color = row["Color"].strip()
        color_num = _color_code(color)
        size = row["Size"].strip()
        dim = row["Dim"].strip()

        if size and dim:
            size_dim = f"{size}/{dim}"
        elif size:
            size_dim = size
        else:
            size_dim = dim

        d_code = f"5-{item_m}-{color_num}"

        output_rows.append([
            "",
            row["Purchase Order"].strip(),
            row["Customer Item"].strip(),
            d_code,
            row["Description"].strip(),
            size_dim,
            _int_value(row["Qty"]),
            float(_money_to_decimal(row["Your Unit Price"])),
            float(_money_to_decimal(row["Amount"])),
            "",
            "",
            5,
            item_m,
            color,
            size,
            dim,
        ])

    return output_rows


def _validate_rows(pdf_rows: List[Dict[str, str]], output_rows: List[List], header: InvoiceHeader) -> List[str]:
    validations = []

    if len(pdf_rows) != len(output_rows):
        raise InvoiceConversionError(f"PDF eilučių skaičius ({len(pdf_rows)}) nesutampa su Excel eilučių skaičiumi ({len(output_rows)}).")

    qty_sum = 0
    amount_sum = Decimal("0.00")

    forbidden_color_words = {
        "GREEN", "COYOTE", "INDIGO", "BLUE", "BROWN", "HTR", "CANOPY", "CAMO",
        "NAVY", "MULTICAM", "KHAKI", "KANGAROO", "BLACK", "CHARCOAL", "TUNDRA",
        "VOLCANIC", "COVERT", "RANGER", "BATTLE", "DARK", "US", "DK", "PACIFIC"
    }

    for idx, (pdf_row, out_row) in enumerate(zip(pdf_rows, output_rows), start=1):
        if out_row[4] != pdf_row["Description"].strip():
            raise InvoiceConversionError(f"{idx} eilutė: Description nesutampa.")
        if out_row[13] != pdf_row["Color"].strip():
            raise InvoiceConversionError(f"{idx} eilutė: Color nesutampa.")
        if out_row[14] != pdf_row["Size"].strip():
            raise InvoiceConversionError(f"{idx} eilutė: Size nesutampa.")
        if out_row[15] != pdf_row["Dim"].strip():
            raise InvoiceConversionError(f"{idx} eilutė: Dim nesutampa.")

        qty = _int_value(pdf_row["Qty"])
        your_unit_price = _money_to_decimal(pdf_row["Your Unit Price"])
        amount = _money_to_decimal(pdf_row["Amount"])

        if qty != out_row[6]:
            raise InvoiceConversionError(f"{idx} eilutė: Qty nesutampa.")
        if Decimal(str(out_row[7])) != your_unit_price:
            raise InvoiceConversionError(f"{idx} eilutė: Your Unit Price nesutampa.")
        if Decimal(str(out_row[8])) != amount:
            raise InvoiceConversionError(f"{idx} eilutė: Amount nesutampa.")

        m_value = out_row[12]
        if str(m_value).endswith("ABR"):
            raise InvoiceConversionError(f"{idx} eilutė: M stulpelyje liko galinis ABR.")

        expected_d = f"5-{m_value}-{_color_code(out_row[13])}"
        if out_row[3] != expected_d:
            raise InvoiceConversionError(f"{idx} eilutė: D stulpelis neteisingas.")

        # Color validation: if Color has spaces, none of its words may leak into Size or Dim.
        color = out_row[13]
        if " " in color:
            size_dim_text = f"{out_row[14]} {out_row[15]}".upper()
            color_words = [part for part in re.split(r"[:\s]+", color.upper()) if part and not part.isdigit()]
            for word in color_words:
                if word in forbidden_color_words and re.search(rf"\b{re.escape(word)}\b", size_dim_text):
                    raise InvoiceConversionError(
                        f"{idx} eilutė: Color žodis {word!r} pateko į Size/Dim."
                    )

        qty_sum += qty
        amount_sum += amount

    invoice_total = _money_to_decimal(header.invoice_total)
    if amount_sum != invoice_total:
        raise InvoiceConversionError(
            f"Amount suma ({amount_sum}) nesutampa su INVOICE TOTAL ({invoice_total})."
        )

    validations.append(f"PDF prekių eilučių: {len(pdf_rows)}")
    validations.append(f"Excel prekių eilučių: {len(output_rows)}")
    validations.append(f"Qty suma: {qty_sum}")
    validations.append(f"Amount suma: {amount_sum}")
    validations.append(f"Invoice Total: {header.invoice_total}")
    validations.append("Color / Size / Dim validacija praėjo.")
    validations.append("M stulpelyje galinis ABR pašalintas.")
    validations.append("D stulpelis sudarytas pagal 5-M-spalvos_kodas.")

    return validations


def _write_xlsx(output_path: Path, header: InvoiceHeader, output_rows: List[List]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.append(["Invoice Number", "Invoice Date", "Customer PO", "Invoice Total"])
    ws.append([header.invoice_number, header.invoice_date, header.customer_po, header.invoice_total])
    ws.append([])
    ws.append(OUTPUT_HEADERS)

    for row in output_rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    table_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = table_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=16):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    for row in range(5, ws.max_row + 1):
        ws[f"G{row}"].number_format = "0"
        ws[f"H{row}"].number_format = "0.00"
        ws[f"I{row}"].number_format = "0.00"
        ws[f"L{row}"].number_format = "0"

    widths = {
        "A": 4, "B": 14, "C": 18, "D": 18, "E": 34, "F": 12, "G": 8, "H": 16,
        "I": 12, "J": 4, "K": 4, "L": 6, "M": 14, "N": 30, "O": 12, "P": 10
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:P{ws.max_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def convert_invoice_pdf(pdf_path: str | Path, output_path: Optional[str | Path] = None) -> ConversionResult:
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise InvoiceConversionError(f"PDF failas nerastas: {pdf_path}")

    if output_path is None:
        output_path = pdf_path.with_suffix(".xlsx")
    output_path = Path(output_path)

    with pdfplumber.open(str(pdf_path)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    header = _extract_header(full_text)
    pdf_rows = _extract_table_rows(pdf_path)
    output_rows = _build_output_rows(pdf_rows)
    validations = _validate_rows(pdf_rows, output_rows, header)
    _write_xlsx(output_path, header, output_rows)

    qty_sum = sum(_int_value(row["Qty"]) for row in pdf_rows)
    amount_sum = sum((_money_to_decimal(row["Amount"]) for row in pdf_rows), Decimal("0.00"))

    return ConversionResult(
        output_path=output_path,
        invoice=header,
        pdf_row_count=len(pdf_rows),
        excel_row_count=len(output_rows),
        qty_sum=qty_sum,
        amount_sum=str(amount_sum),
        validations=validations,
    )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Convert 5.11 invoice PDF to validated Excel.")
    parser.add_argument("pdf", help="Input PDF path")
    parser.add_argument("-o", "--output", help="Output XLSX path", default=None)
    args = parser.parse_args()

    result = convert_invoice_pdf(args.pdf, args.output)
    print(f"OK: {result.output_path}")
    print("\n".join(result.validations))
